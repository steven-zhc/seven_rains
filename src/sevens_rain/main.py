"""Main module with new modular architecture."""

import openpyxl
from pathlib import Path
from datetime import datetime, timedelta, date
from typing import List, Dict, Optional

# New architecture imports
from .models import DayType, WeekPlan
from .scheduler import WeekScheduler
from .storage import PlanStorage
from .excel_generator import ExcelGenerator
from .rules import DEFAULT_RULES

# Employee names - extracted from sample.xls file
EMPLOYEES = [
    "姚强",
    "钱国祥", 
    "包汀池",
    "孙震", 
    "赵兵",
    "夏银龙",
    "张尧"
]

# Global instances for new architecture
storage = PlanStorage("plan.json")
scheduler = WeekScheduler(EMPLOYEES, DEFAULT_RULES)
excel_generator = ExcelGenerator(EMPLOYEES, storage)


def extract_real_employee_names(file_path: str) -> List[str]:
    """Extract real employee names from sample Excel file using openpyxl."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        names = []
        
        # Names are in column 2 (B), starting from row 5 (after header rows)
        for row in range(5, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=2).value  # Column B contains names
            if cell_value is None:
                # Stop when we hit empty cells (end of employee list)
                break
            
            name = str(cell_value).strip()
            if name:
                names.append(name)
                
        workbook.close()
        return names
    except Exception as e:
        print(f"Error extracting employee names: {e}")
        return []


def analyze_sample_excel(file_path: str) -> dict:
    """Analyze the sample Excel file to understand its structure."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        analysis = {
            'sheets': workbook.sheetnames,
            'sheet_info': {}
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            analysis['sheet_info'][sheet_name] = {
                'shape': (sheet.max_row, sheet.max_column),
                'dimensions': f'{sheet.min_row}-{sheet.max_row} rows, {sheet.min_column}-{sheet.max_column} cols'
            }
        
        workbook.close()
        return analysis
    except Exception as e:
        return {'error': str(e)}


def generate_month_schedule_new(year: int, month: int) -> str:
    """Generate month schedule using new architecture."""
    return excel_generator.generate_month_schedule(year, month)


def get_employee_summary_new(year: int, month: int) -> Dict[str, Dict[str, int]]:
    """Get employee statistics using new architecture."""
    # Get month weeks from storage
    month_weeks = storage.get_month_weeks(year, month)
    
    if not month_weeks:
        # Generate weeks if not exist
        excel_generator.generate_month_schedule(year, month)
        month_weeks = storage.get_month_weeks(year, month)
    
    # Calculate statistics
    employee_stats = {emp: {"听": 0, "休": 0, "白": 0} for emp in EMPLOYEES}
    
    for week in month_weeks:
        for day in range(7):
            for employee in EMPLOYEES:
                day_type = week.get_assignment(day, employee)
                if day_type:
                    employee_stats[employee][day_type.value] += 1
    
    return employee_stats




def generate_excel(year: int = 2025, month: int = 9, output_path: Optional[str] = None) -> None:
    """Generate Excel file using new architecture."""
    if output_path:
        # Custom output path provided
        excel_generator.generate_month_schedule(year, month, output_path)
    else:
        # Use default naming
        excel_generator.generate_month_schedule(year, month)


def main(target_year: int = None, target_month: int = None) -> None:
    """Main CLI entry point with new architecture."""
    print("Seven Rain - Excel Schedule Generator (New Architecture)")
    print("🎯 Features: Modular rules, cross-month weeks, JSON metadata")
    
    # Use current date if no target specified
    if target_year is None or target_month is None:
        from datetime import date
        today = date.today()
        target_year = target_year or today.year
        target_month = target_month or today.month
    
    # Try to load real employee names from sample file
    sample_path = "sample.xls"
    real_employees = []
    if Path(sample_path).exists():
        real_employees = extract_real_employee_names(sample_path)
        if real_employees:
            print(f"Loaded {len(real_employees)} employees from {sample_path}")
            print(f"Employees: {', '.join(real_employees)}")
            
            # Update global instances with real employees
            global EMPLOYEES, scheduler, excel_generator
            EMPLOYEES = real_employees
            scheduler = WeekScheduler(EMPLOYEES, DEFAULT_RULES)
            excel_generator = ExcelGenerator(EMPLOYEES, storage)
        else:
            print(f"Using hardcoded employees: {', '.join(EMPLOYEES)}")
    else:
        print(f"Sample file not found, using hardcoded employees: {', '.join(EMPLOYEES)}")
    
    # Show active rules
    print(f"\n📋 Active Rules ({len(scheduler.rules)}):")
    for rule in scheduler.rules:
        print(f"  • {rule.get_name()} (Priority: {rule.get_priority()})")
    
    # Generate schedule for target month
    month_names = ["", "January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    print(f"\n🔄 Generating {month_names[target_month]} {target_year} schedule...")
    output_file = generate_month_schedule_new(target_year, target_month)
    
    # Get employee summary using new architecture
    summary = get_employee_summary_new(target_year, target_month)
    
    print(f"\n📊 Employee Summary ({month_names[target_month]} {target_year}):")
    for emp in EMPLOYEES:
        stats = summary[emp]
        print(f"  {emp}: 值班{stats['听']}天, 休息{stats['休']}天, 工作{stats['白']}天")
    
    # Storage statistics
    storage_stats = storage.get_statistics()
    print(f"\n💾 Storage Info:")
    print(f"  Total weeks stored: {storage_stats['total_weeks']}")
    if storage_stats['total_weeks'] > 0:
        print(f"  Date range: {storage_stats['earliest_week']} to {storage_stats['latest_week']}")
    
    print(f"\n✅ Excel generated: {output_file}")
    
    # Also analyze sample file if exists
    if Path(sample_path).exists():
        print("\n--- Sample Analysis ---")
        analysis = analyze_sample_excel(sample_path)
        if 'error' not in analysis:
            print(f"Sample has {len(analysis.get('sheets', []))} sheets")
            for sheet, info in analysis.get('sheet_info', {}).items():
                print(f"Sheet '{sheet}': {info['shape']} (rows x cols)")
        else:
            print(f"Error: {analysis['error']}")


if __name__ == "__main__":
    import sys
    
    target_year = None
    target_month = None
    
    # Parse command line arguments
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if '-' in arg:  # Format: 2025-10
            year_str, month_str = arg.split('-')
            target_year = int(year_str)
            target_month = int(month_str)
        else:  # Just year
            target_year = int(arg)
    
    main(target_year, target_month)